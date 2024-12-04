from tkinter import *
from tkinter import messagebox
import openpyxl

root = Tk()
root.title("Empirical to Molecular")

periodictableelementsymbols=['H','He','Li','Be','B','C','N','O','F','Ne','Na','Mg','Al','Si','P','S','Cl','Ar','K','Ca','Sc','Ti','V','Cr','Mn','Fe','Co','Ni','Cu','Zn','Ga','Ge','As','Se','Br','Kr','Rb','Sr','Y','Zr','Nb','Mo','Tc','Ru','Rh','Pd','Ag','Cd','In','Sn','Sb','Te','I','Xe','Cs','Ba','La','Ce','Pr','Nd','Pm','Sm','Eu','Gd','Tb','Dy','Ho','Er','Tm','Yb','Lu','Hf','Ta','W','Re','Os','Ir','Pt','Au','Hg','Tl','Pb','Bi','Po','At','Rn','Fr','Ra','Ac','Pa','Th','Np','U','Am','Pu','Cm','Bk','Cf','Es','Fm','Md','No','Rf','Lr','Db','Bh','Sg','Mt','Rg','Hs']
periodictableatomicnumbers=[1,4,7,9,11,12,14,16,19,20,23,24,27,28,31,32,35.5,40,39,40,45,48,51,52,55,56,59,59,64,65,70,73,75,79,80,84,85,88,89,91,93,96,98,101,103,106,108,112,115,119,122,128,127,131,133,137,139,140,141,144,145,150,152,157,159,162,165,167,169,173,175,178,181,184,186,190,192,195,197,201,204,207,209,209,210,222,223,226,227,231,232,237,238,243,244,247,247,251,252,257,258,259,261,262,262,264,266,268,272,277]
periodictableelements=['Hydrogen','Helium','Lithium','Beryllium','Boron','Carbon','Nitrogen','Oxygen','Fluorine','Neon','Sodium','Magnesium','Aluminum','Silicon','Phosphorus','Sulphur','Chlorine','Argon','Potassium','Calcium','Scandium','Titanium','Vanadium','Chromium','Manganese','Iron','Cobalt','Nickel','Copper','Zinc','Gallium','Germanium','Arsenic','Selenium','Bromine','Krypton','Rubidium','Strontium','Yttrium','Zirconium','Neobium','Molybdenum','Technetium','Rutinium','Rhodium','Palladium','Silver','Cadmium','Indium','Tin','Antimony','Iodine','Tellurium','Zenon','Caesium','Barium','Lanthoanum','Cerium','Praseodymium','Neodymium','Promethium','Samerium','Europium','Gadolinium','Turbium','Dysprosium','Holmium','Herbium','Thulium','Ytterbium','Lutetium','Halfnium','Tantalum','Tungsten','Rhenium','Osmium','Iridium','Platinum','Gold','Mercury','Thallium','Lead','Bismuth','Polonium','Astetine','Radon','Francium','Radium','Actinium','Protactinium','Thorium','Neptunium','Uranium','Americium','Plutonium','Curium','Berkilium','Californium','Einsteinium','Fermium','Mendelevium','Nobelium','Rutherforduim','Lorencium','Dubnium','Bohrium','Seaborgium','Meitnerium','Roentgenium','Hassium']
global empiricalelements            #initialising arrays and variables required in the code
empiricalelements=[]
global numberofmolecules
numberofmolecules=[]
total=0
finishcertify=0
mass=0

Elementlabel = Label(root, text="Element:")      #creating input boxes and labels used as part of the GUI interface
Elementlabel.grid(row=1,column=1)

element = Entry(root,width=30,borderwidth=2)
element.grid(row=1,column=2)

molabel = Label(root, text="Molecules:")
molabel.grid(row=2,column=1)

nom = Entry(root,width=30,borderwidth=2)
nom.grid(row=2,column=2)

malabel = Label(root, text="Total mass:")
malabel.grid(row=3,column=1)

molecularmass = Entry(root,width=30,borderwidth=2)
molecularmass.grid(row=3,column=2)

Resultlabel = Label(root, text="Results:")
Resultlabel.grid(row=4,column=1)

frame = Frame(root)
frame.grid(row=4,column=2)

def starting():
    global total
    g=-1
    elemento=element.get()
    for b in range (0,110):
        if (periodictableelementsymbols[b]==str(elemento)) or (periodictableelements[b]==str(elemento)):     #Checks to see if element entered exists
            g=b
    nm=nom.get()
    if (len(nm)==0):
        messagebox.showinfo("Error","No molecule input")   #Error box shown if presence check shows no % or mass input
    elif (g>-1) and (int(nm)>0):
        empiricalelements.append(str(elemento))
        numberofmolecules.append(int(nm))
        total=total+(int(nm) * periodictableatomicnumbers[g])
        element.delete(0,END)
        nom.delete(0,END)
    elif (g>-1):
        messagebox.showinfo("Error","Number of molecules out of acceptable range")      #Error box shown if mass or % input is negative
    else:
        messagebox.showinfo("Error","Element does not exist")       #Error box shown if element input is not acceptable
        
def finish():
    global finishcertify          #Initialising more variabes required in the code
    global mass
    global molmass
    mass=molecularmass.get()      #Taking the mass from the input boxes
    if (len(mass)!=0):
        molmass=molecularmass.get()
        molmass=float(molmass)/total  #Dividing the total mass by the total mass of the elements to get a multiple
        for i in range (len(empiricalelements)):
            myLabel = Label(frame, text= empiricalelements[i] + ": " + str((molmass*numberofmolecules[i]))) #Creates a frame to output the results for each element one-by-one
            myLabel.grid(row=i+8,column=1)                
        if (len(empiricalelements)>0):
            finishcertify=1
        molecularmass.delete(0,END)
    else:
        messagebox.showinfo("Error","No mass input")
def resetting():
    global finishcertify                        #Resets all array and varaibles while clearing out any data in boxes
    global total
    if (finishcertify==1):
        for label in frame.winfo_children():    #Deletes all the children in the frame
            label.destroy()
    empiricalelements.clear()
    numberofmolecules.clear()
    total=0
    finishcertify=0
    element.delete(0,END)
    nom.delete(0,END)
    molecularmass.delete(0,END)

def storedata():
    wb=openpyxl.load_workbook('EftoMf.xlsx')                #Opens excel file named EftoMf
    sheet=wb.active
    x=sheet.max_row                                         #Looks for last row used in file so new data can be placed after that
    x=x+1
    a='A'+str(x)
    sheet[a] = mass
    
    for i in range (x,x+len(empiricalelements)): 
        b='B'+str(i)
        c='C'+str(i)
        d='D'+str(i)
        e='E'+str(i)
        sheet[b] = empiricalelements[i-x]                   #Adding the data to the sheet
        sheet[d] = empiricalelements[i-x]
        sheet[c] = numberofmolecules[i-x]
        sheet[e] = (molmass*numberofmolecules[i-x])
    y=sheet.max_row
    y=y+1
    f='A'+str(y)                                            #Separating the data from the next set of input data using '-'
    g='B'+str(y)
    h='C'+str(y)
    i='D'+str(y)
    j='E'+str(y)
    sheet[f] = "-"
    sheet[g] = "-"
    sheet[h] = "-"
    sheet[i] = "-"
    sheet[j] = "-"
    wb.save('EftoMf.xlsx')
    
confirm = Button(root, text="Input Element", command=starting)    #Creating buttons for the GUI interface
confirm.grid(row=1,column=3)

finishing = Button(root, text="Finish", command=finish)
finishing.grid(row=2,column=3)

reset = Button(root, text="Reset", command=resetting)
reset.grid(row=3,column=3)

store = Button(root, text="Store", command=storedata)
store.grid(row=4,column=3)

root.mainloop()

