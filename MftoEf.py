from tkinter import *
from tkinter import messagebox
import openpyxl

root = Tk()
root.title("Molecular to Empirical")

periodictableelementsymbols=['H','He','Li','Be','B','C','N','O','F','Ne','Na','Mg','Al','Si','P','S','Cl','Ar','K','Ca','Sc','Ti','V','Cr','Mn','Fe','Co','Ni','Cu','Zn','Ga','Ge','As','Se','Br','Kr','Rb','Sr','Y','Zr','Nb','Mo','Tc','Ru','Rh','Pd','Ag','Cd','In','Sn','Sb','Te','I','Xe','Cs','Ba','La','Ce','Pr','Nd','Pm','Sm','Eu','Gd','Tb','Dy','Ho','Er','Tm','Yb','Lu','Hf','Ta','W','Re','Os','Ir','Pt','Au','Hg','Tl','Pb','Bi','Po','At','Rn','Fr','Ra','Ac','Pa','Th','Np','U','Am','Pu','Cm','Bk','Cf','Es','Fm','Md','No','Rf','Lr','Db','Bh','Sg','Mt','Rg','Hs']
periodictableatomicnumbers=[1,4,7,9,11,12,14,16,19,20,23,24,27,28,31,32,35.5,40,39,40,45,48,51,52,55,56,59,59,64,65,70,73,75,79,80,84,85,88,89,91,93,96,98,101,103,106,108,112,115,119,122,128,127,131,133,137,139,140,141,144,145,150,152,157,159,162,165,167,169,173,175,178,181,184,186,190,192,195,197,201,204,207,209,209,210,222,223,226,227,231,232,237,238,243,244,247,247,251,252,257,258,259,261,262,262,264,266,268,272,277]
periodictableelements=['Hydrogen','Helium','Lithium','Beryllium','Boron','Carbon','Nitrogen','Oxygen','Fluorine','Neon','Sodium','Magnesium','Aluminum','Silicon','Phosphorus','Sulphur','Chlorine','Argon','Potassium','Calcium','Scandium','Titanium','Vanadium','Chromium','Manganese','Iron','Cobalt','Nickel','Copper','Zinc','Gallium','Germanium','Arsenic','Selenium','Bromine','Krypton','Rubidium','Strontium','Yttrium','Zirconium','Neobium','Molybdenum','Technetium','Rutinium','Rhodium','Palladium','Silver','Cadmium','Indium','Tin','Antimony','Iodine','Tellurium','Zenon','Caesium','Barium','Lanthoanum','Cerium','Praseodymium','Neodymium','Promethium','Samerium','Europium','Gadolinium','Turbium','Dysprosium','Holmium','Herbium','Thulium','Ytterbium','Lutetium','Halfnium','Tantalum','Tungsten','Rhenium','Osmium','Iridium','Platinum','Gold','Mercury','Thallium','Lead','Bismuth','Polonium','Astetine','Radon','Francium','Radium','Actinium','Protactinium','Thorium','Neptunium','Uranium','Americium','Plutonium','Curium','Berkilium','Californium','Einsteinium','Fermium','Mendelevium','Nobelium','Rutherforduim','Lorencium','Dubnium','Bohrium','Seaborgium','Meitnerium','Roentgenium','Hassium']
global elements              #initialising arrays and variables required in the code
elements=[]
global percetagesormasses
percentagesormasse=[]
percentagesormasses3=[]
low=0
finishcertify=0

Elementlabel = Label(root, text="Element:")             #creating input boxes and labels used as part of the GUI interface 
Elementlabel.grid(row=1,column=1)

Pormlabel = Label(root, text="% or Mass:")
Pormlabel.grid(row=2,column=1)

Resultlabel = Label(root, text="Results:")
Resultlabel.grid(row=3,column=1)

element = Entry(root,width=30,borderwidth=2)
element.grid(row=1,column=2)

porm = Entry(root,width=30,borderwidth=2)
porm.grid(row=2,column=2)

frame = Frame(root)
frame.grid(row=3,column=2)

def starting():
    global low
    global percentagesormasses3
    g=-1
    elemento=element.get()
    for b in range (0,110):    
        if (periodictableelementsymbols[b]==str(elemento)) or (periodictableelements[b]==str(elemento)): #Checks to see if element entered exists
            g=b
    pm=porm.get()
    if (len(pm)==0):
        messagebox.showinfo("Error","No % or mass input")    #Error box shown if presence check shows no % or mass input 
    elif (g>-1) and (float(pm)>0):
        elements.append(str(elemento))
        element.delete(0,END)    
        percentagesormasses3.append(float(pm))
        percentagesormasse.append(float(pm)/periodictableatomicnumbers[g])  #Divides the mass of the 
        if (low==0):
            low=(float(pm)/periodictableatomicnumbers[g])                           
        elif (low>(float(pm)/periodictableatomicnumbers[g])):
            low=(float(pm)/periodictableatomicnumbers[g])
        porm.delete(0,END)
    elif (g>-1):
        messagebox.showinfo("Error","Mass or % is out of acceptable range")  #Error box shown if mass or % input is negative
    else:
        messagebox.showinfo("Error","Element does not exist")  #Error box shown if element input is not acceptable

def finish():
    global finishcertify          #Initialising more variabes required in the code
    global percentagesormasses2
    percentagesormasses=percentagesormasse
    percentagesormasses2=[]
    ranges=0
    multiplication=1
    for y in range (len(elements)):
        percentagesormasses2.append(percentagesormasses[y]/low)  #Converting each element into their empirical form by using ratios
    for a in range (len(elements)):
        if (percentagesormasses2[a]-round(percentagesormasses2[a])>0.13) or (percentagesormasses2[a]-round(percentagesormasses2[a])<-0.13):    #ACCORDING TO THE SET PARAMETERS ALLOWED BY THE TEACHER, THE 0.13 RANGE WILL CHANGE
            ranges=ranges+1
    if (ranges!=0):
        while (1>0):
            ranges=0
            multiplication=multiplication+1
            percentagesormasses2=[]
            for f in range (len(elements)):
                percentagesormasses2[f].append(percentagesormasses[f]*multiplication)
            for a in range (len(elements)):
                if (percentagesormasses2[a]-round(percentagesormasses2[a])>0.13) or (percentagesormasses2[a]-round(percentagesormasses2[a])<-0.13):     #ACCORDING TO THE SET PARAMETERS ALLOWED BY THE TEACHER, THE 0.13 RANGE WILL CHANGE
                    ranges=ranges+1
            if (ranges==0):
                break
    for z in range (0,len(percentagesormasses2)):
        myLabel = Label(frame, text= elements[z] + ": " + str(percentagesormasses2[z])) #Creates a frame to output the results for each element one-by-one
        myLabel.grid(row=z,column=3)
    if (len(percentagesormasses2)>0):
        finishcertify=1
    
def resetting():
    global finishcertify
    if (finishcertify==1):
        for label in frame.winfo_children():   #Resets all array and varaibles while clearing out any data in boxes
            label.destroy()
    elements.clear()
    percentagesormasse.clear()                 #Deletes all the children in the frame
    finishcertify=0
    element.delete(0,END)
    porm.delete(0,END)
    percentagesormasses3.clear()

def storedata():
    global percentagesormasses3
    wb=openpyxl.load_workbook('MftoEf.xlsx')    #Opens excel file named MftoEf
    sheet=wb.active
    x=sheet.max_row                             #Looks for last row used in file so new data can be placed after that
    for i in range (x+1,x+1+len(elements)):
        a='A'+str(i)
        b='B'+str(i)
        c='C'+str(i)
        d='D'+str(i)
        sheet[a] = elements[i-(x+1)]            #Adding the data to the sheet
        sheet[c] = elements[i-(x+1)]
        sheet[b] = percentagesormasses3[i-(x+1)]
        sheet[d] = percentagesormasses2[i-(x+1)]
    y=sheet.max_row
    e='A'+str(y+1)                              #Separating the data from the next set of input data using '-'
    f='B'+str(y+1)
    g='C'+str(y+1)
    h='D'+str(y+1)
    sheet[e] = "-"
    sheet[f] = "-"
    sheet[g] = "-"
    sheet[h] = "-"
    wb.save('MftoEf.xlsx')
    
confirm = Button(root, text="Input Element", command=starting) #Creating buttons for the GUI interface
confirm.grid(row=1,column=4)

finishing = Button(root, text="Finish", command=finish)
finishing.grid(row=2,column=4)

reset = Button(root, text="Reset", command=resetting)
reset.grid(row=3,column=4)

store = Button(root, text="Store", command=storedata)
store.grid(row=4,column=4)

root.mainloop()

